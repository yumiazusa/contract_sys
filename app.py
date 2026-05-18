from flask import Flask, render_template, request, jsonify, send_file, redirect, url_for, session, after_this_request
from flask_sqlalchemy import SQLAlchemy
from sqlalchemy import or_, func
from sqlalchemy.exc import IntegrityError
from datetime import datetime, timedelta, date
from werkzeug.security import generate_password_hash, check_password_hash
import openpyxl
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Font, Alignment, PatternFill
import os
import tempfile

try:
    from dotenv import load_dotenv
except ImportError:
    def load_dotenv():
        return False

load_dotenv()

DEFAULT_DB_URI = 'mysql+pymysql://ProjectDB:4100282Ly%40@47.108.254.13/projectdb'
DEFAULT_PAGE_SIZE = 50
MAX_PAGE_SIZE = 200
EXPORT_BATCH_SIZE = 200


def get_int_env(name, default, minimum=None, maximum=None):
    raw_value = os.getenv(name)
    try:
        value = int(raw_value) if raw_value is not None else default
    except (TypeError, ValueError):
        value = default

    if minimum is not None:
        value = max(minimum, value)
    if maximum is not None:
        value = min(maximum, value)
    return value


def get_bool_env(name, default=False):
    raw_value = os.getenv(name)
    if raw_value is None:
        return default
    return raw_value.strip().lower() in {'1', 'true', 'yes', 'on'}


def normalize_text(value):
    return (value or '').strip()


database_uri = os.getenv('DATABASE_URL', DEFAULT_DB_URI)
engine_options = {'pool_pre_ping': True}
if not database_uri.startswith('sqlite'):
    engine_options.update({
        'pool_recycle': get_int_env('DB_POOL_RECYCLE', 1800, minimum=60),
        'pool_size': get_int_env('DB_POOL_SIZE', 3, minimum=1, maximum=10),
        'max_overflow': get_int_env('DB_MAX_OVERFLOW', 2, minimum=0, maximum=10),
        'pool_timeout': get_int_env('DB_POOL_TIMEOUT', 30, minimum=5, maximum=120)
    })

app = Flask(__name__)
app.secret_key = os.getenv('SECRET_KEY', 'your-secret-key-change-this-in-production')
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(days=30)

# 数据库配置
# 临时切换到本地SQLite数据库以解决远程连接权限问题
#app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///contract.db'
#原MySQL配置（如需恢复远程连接，请取消注释下方代码并注释掉上方SQLite配置）
app.config['SQLALCHEMY_DATABASE_URI'] = database_uri
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SQLALCHEMY_ENGINE_OPTIONS'] = engine_options
db = SQLAlchemy(app)


# 数据库模型
class User(db.Model):
    __tablename__ = 'users'
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(50), unique=True, nullable=False)
    password = db.Column(db.String(255), nullable=False)
    realname = db.Column(db.String(100))
    department = db.Column(db.String(50))
    created_date = db.Column(db.DateTime, default=datetime.now)

class Contract(db.Model):
    __tablename__ = 'contracts'

    id = db.Column(db.Integer, primary_key=True)
    contract_no = db.Column(db.String(50), unique=True, nullable=False)
    contract_name = db.Column(db.String(200), nullable=False)
    project_no = db.Column(db.String(500))
    contract_type = db.Column(db.String(20), nullable=False)
    platform = db.Column(db.String(10), nullable=False)
    contract_amount = db.Column(db.Numeric(15, 2))
    sign_date = db.Column(db.Date)
    company_name = db.Column(db.String(200), nullable=False)
    contact_phone = db.Column(db.String(50), nullable=False)
    corporate_principal = db.Column(db.String(100), nullable=False)
    department = db.Column(db.String(50), nullable=False)
    payment_terms = db.Column(db.Text)
    original_contract_no = db.Column(db.String(50))
    original_contract_name = db.Column(db.String(200))
    remarks = db.Column(db.Text)
    created_at = db.Column(db.DateTime, default=datetime.now)
    updated_at = db.Column(db.DateTime, default=datetime.now, onupdate=datetime.now)
    executive_partner = db.Column(db.String(255))
    filler = db.Column(db.String(255))
    status = db.Column(db.String(20), default='active')  # active, invalid


def serialize_contract(contract):
    return {
        'id': contract.id,
        'contract_no': contract.contract_no,
        'contract_name': contract.contract_name,
        'project_no': contract.project_no,
        'contract_type': contract.contract_type,
        'platform': contract.platform,
        'contract_amount': str(contract.contract_amount) if contract.contract_amount else '',
        'sign_date': contract.sign_date.strftime('%Y-%m-%d') if contract.sign_date else '',
        'company_name': contract.company_name,
        'contact_phone': contract.contact_phone,
        'corporate_principal': contract.corporate_principal,
        'department': contract.department,
        'payment_terms': contract.payment_terms,
        'original_contract_no': contract.original_contract_no,
        'original_contract_name': contract.original_contract_name,
        'remarks': contract.remarks,
        'created_at': contract.created_at.strftime('%Y-%m-%d %H:%M:%S') if contract.created_at else '',
        'updated_at': contract.updated_at.strftime('%Y-%m-%d %H:%M:%S') if contract.updated_at else '',
        'executive_partner': contract.executive_partner,
        'filler': contract.filler,
        'status': contract.status
    }


def get_payload_value(data, key):
    value = data.get(key)
    return value if value != '' else None


def apply_contract_filters(query, params):
    keyword = normalize_text(params.get('keyword'))
    contract_type = normalize_text(params.get('contract_type'))
    platform = normalize_text(params.get('platform'))
    filler = normalize_text(params.get('filler'))
    exec_partner = normalize_text(params.get('executive_partner'))
    department = normalize_text(params.get('department'))
    status = normalize_text(params.get('status'))

    if keyword:
        like_keyword = f'%{keyword}%'
        query = query.filter(or_(
            Contract.contract_no.like(like_keyword),
            Contract.contract_name.like(like_keyword),
            Contract.company_name.like(like_keyword),
            Contract.project_no.like(like_keyword)
        ))
    if contract_type:
        query = query.filter(Contract.contract_type == contract_type)
    if platform:
        query = query.filter(Contract.platform == platform)
    if filler:
        query = query.filter(Contract.filler.like(f'%{filler}%'))
    if exec_partner:
        query = query.filter(Contract.executive_partner.like(f'%{exec_partner}%'))
    if department:
        query = query.filter(Contract.department == department)
    if status:
        query = query.filter(Contract.status == status)

    return query


def get_pagination_params():
    page = get_int_env('DEFAULT_PAGE', 1, minimum=1)
    page_size = DEFAULT_PAGE_SIZE

    request_page = request.args.get('page')
    request_page_size = request.args.get('page_size')

    if request_page is not None:
        try:
            page = max(1, int(request_page))
        except (TypeError, ValueError):
            page = 1

    if request_page_size is not None:
        try:
            page_size = min(MAX_PAGE_SIZE, max(1, int(request_page_size)))
        except (TypeError, ValueError):
            page_size = DEFAULT_PAGE_SIZE

    return page, page_size


def get_contract_summary():
    today = date.today()
    today_start = datetime.combine(today, datetime.min.time())
    month_start = datetime(today.year, today.month, 1)

    total_contracts = db.session.query(func.count(Contract.id)).scalar() or 0
    today_contracts = db.session.query(func.count(Contract.id)).filter(Contract.created_at >= today_start).scalar() or 0
    month_contracts = db.session.query(func.count(Contract.id)).filter(Contract.created_at >= month_start).scalar() or 0
    total_amount = db.session.query(func.coalesce(func.sum(Contract.contract_amount), 0)).filter(
        Contract.status != 'invalid'
    ).scalar() or 0

    return {
        'total_contracts': int(total_contracts),
        'today_contracts': int(today_contracts),
        'month_contracts': int(month_contracts),
        'total_amount': float(total_amount)
    }


# 生成合同编号
def generate_contract_no(contract_type, platform):
    prefix = 'KJ' if contract_type == '框架合同' else 'HT'
    platform_code = 'JQ' if platform == '金乾' else 'JC'
    current_year = datetime.now().year

    # 查询当前平台和类型的最大流水号
    # 使用 like 查询确保匹配正确的前缀（年份可能会变）
    year_prefix = f"{prefix}{current_year}{platform_code}"
    last_contract = Contract.query.filter(
        Contract.contract_no.like(f"{year_prefix}%")
    ).order_by(Contract.contract_no.desc()).first()

    if last_contract:
        try:
            last_no = int(last_contract.contract_no[-4:])
            new_no = last_no + 1
        except ValueError:
            new_no = 1
    else:
        new_no = 1
    
    # 最终生成
    final_no = f"{year_prefix}{new_no:04d}"
    
    # 双重保险：检查是否存在，如果存在则递增直到唯一
    while Contract.query.filter_by(contract_no=final_no).first():
        new_no += 1
        final_no = f"{year_prefix}{new_no:04d}"

    return final_no


# 登录页面
@app.route('/')
def index():
    if 'user' in session:
        return redirect(url_for('dashboard'))
    return render_template('login.html')


# 登录验证
@app.route('/login', methods=['POST'])
def login():
    username = request.form.get('username')
    password = request.form.get('password')

    user = User.query.filter_by(username=username).first()
    
    # 验证逻辑：
    # 1. 如果数据库中存在该用户
    # 2. 检查密码是否匹配（支持明文比对或 Hash 比对，优先 Hash）
    # 3. 如果是 admin/admin123，且数据库中没有该用户，我们稍后通过 init_admin 确保其存在
    
    is_valid = False
    if user:
        if check_password_hash(user.password, password):
            is_valid = True
        elif user.password == password: # 兼容明文存储的情况
            is_valid = True
            
    if is_valid:
        session.permanent = True
        session['user'] = user.username
        session['realname'] = user.realname
        return redirect(url_for('dashboard'))
        
    return render_template('login.html', error='用户名或密码错误')


# 初始化数据库和管理员账号
def init_admin():
    with app.app_context():
        admin = User.query.filter_by(username='admin').first()
        if not admin:
            admin = User(
                username='admin',
                password=generate_password_hash('admin123'),
                realname='系统管理员',
                department='管理部'
            )
            db.session.add(admin)
            db.session.commit()
            print("Default admin user created.")

# 在应用启动时尝试初始化管理员
# init_admin()


# 退出登录
@app.route('/logout')
def logout():
    session.pop('user', None)
    session.pop('realname', None)
    return redirect(url_for('index'))


# 主页面
@app.route('/dashboard')
def dashboard():
    if 'user' not in session:
        return redirect(url_for('index'))
    return render_template('dashboard.html')


# 获取合同列表
@app.route('/api/contracts', methods=['GET'])
def get_contracts():
    page, page_size = get_pagination_params()
    query = apply_contract_filters(Contract.query, request.args)
    pagination = query.order_by(Contract.created_at.desc(), Contract.id.desc()).paginate(
        page=page,
        per_page=page_size,
        error_out=False
    )

    return jsonify({
        'items': [serialize_contract(contract) for contract in pagination.items],
        'pagination': {
            'page': pagination.page,
            'page_size': page_size,
            'pages': pagination.pages,
            'total': pagination.total
        }
    })


@app.route('/api/contracts/summary', methods=['GET'])
def get_contracts_summary():
    return jsonify(get_contract_summary())


@app.route('/api/contracts/<int:contract_id>', methods=['GET'])
def get_contract(contract_id):
    contract = Contract.query.get_or_404(contract_id)
    return jsonify(serialize_contract(contract))


# 获取筛选选项
@app.route('/api/contracts/filter_options', methods=['GET'])
def get_filter_options():
    partners = db.session.query(Contract.executive_partner).distinct().filter(Contract.executive_partner.isnot(None)).all()
    fillers = db.session.query(Contract.filler).distinct().filter(Contract.filler.isnot(None)).all()
    
    return jsonify({
        'executive_partners': sorted([p[0] for p in partners if p[0]]),
        'fillers': sorted([f[0] for f in fillers if f[0]])
    })


# 检查重复合同
@app.route('/api/contracts/check_duplicate', methods=['POST'])
def check_duplicate():
    data = request.json
    company_name = data.get('company_name')
    contract_name = data.get('contract_name')
    contract_amount = data.get('contract_amount')
    edit_id = data.get('edit_id')

    if not contract_amount:
        return jsonify({'duplicates': []})
        
    try:
        # Convert to float for query, though DB is Numeric
        amount_val = float(contract_amount)
    except ValueError:
        return jsonify({'duplicates': []})

    # 查询逻辑：(单位名称 + 合同金额) OR (合同名称 + 合同金额)
    # 且状态不为作废 (假设作废的不算重复？或者都算？需求说是"历史合同列表"，通常包含所有，但作废的可能不算冲突。
    # 既然是防止重复提交，通常是防止有效合同重复。我先只查 active 的，或者全部。)
    # 需求说 "历史合同列表"，没说排除作废。但一般业务逻辑排除 invalid。
    # 既然没有明确说排除，我先只查 active。
    
    query = Contract.query.filter(
        Contract.status != 'invalid',
        or_(
            (Contract.company_name == company_name) & (Contract.contract_amount == amount_val),
            (Contract.contract_name == contract_name) & (Contract.contract_amount == amount_val)
        )
    )

    if edit_id:
        query = query.filter(Contract.id != edit_id)

    duplicates = query.order_by(Contract.created_at.desc(), Contract.id.desc()).all()

    return jsonify({
        'duplicates': [serialize_contract(contract) for contract in duplicates]
    })


# 创建合同
@app.route('/api/contracts', methods=['POST'])
def create_contract():
    data = request.json
    
    # 记录强制提交日志
    if data.get('force_submit'):
        print(f"[AUDIT] Force submit duplicate contract: {data.get('contract_name')} by {session.get('realname', 'unknown')}")

    for _ in range(5):
        contract_no = generate_contract_no(data['contract_type'], data['platform'])
        contract = Contract(
            contract_no=contract_no,
            contract_name=data['contract_name'],
            project_no=get_payload_value(data, 'project_no'),
            contract_type=data['contract_type'],
            platform=data['platform'],
            contract_amount=get_payload_value(data, 'contract_amount'),
            sign_date=datetime.strptime(data['sign_date'], '%Y-%m-%d') if get_payload_value(data, 'sign_date') else None,
            company_name=data['company_name'],
            contact_phone=data['contact_phone'],
            corporate_principal=data['corporate_principal'],
            department=data['department'],
            payment_terms=get_payload_value(data, 'payment_terms'),
            original_contract_no=get_payload_value(data, 'original_contract_no'),
            original_contract_name=get_payload_value(data, 'original_contract_name'),
            remarks=get_payload_value(data, 'remarks'),
            executive_partner=get_payload_value(data, 'executive_partner'),
            filler=get_payload_value(data, 'filler'),
            status='active'
        )

        db.session.add(contract)
        try:
            db.session.commit()
            return jsonify({'success': True, 'contract_no': contract_no, 'id': contract.id})
        except IntegrityError:
            db.session.rollback()

    return jsonify({'success': False, 'message': '合同编号生成冲突，请稍后重试'}), 409


# 更新合同
@app.route('/api/contracts/<int:contract_id>', methods=['PUT'])
def update_contract(contract_id):
    contract = Contract.query.get_or_404(contract_id)
    if contract.status == 'invalid':
        return jsonify({'success': False, 'message': '已作废的合同不可编辑'}), 400
        
    data = request.json

    if data.get('force_submit'):
        print(f"[AUDIT] Force update duplicate contract: {data.get('contract_name')} by {session.get('realname', 'unknown')}")

    contract.contract_name = data['contract_name']
    contract.contract_type = data['contract_type']
    contract.platform = data['platform']
    contract.project_no = get_payload_value(data, 'project_no')
    contract.contract_amount = get_payload_value(data, 'contract_amount')
    contract.sign_date = datetime.strptime(data['sign_date'], '%Y-%m-%d') if get_payload_value(data, 'sign_date') else None
    contract.company_name = data['company_name']
    contract.contact_phone = data['contact_phone']
    contract.corporate_principal = data['corporate_principal']
    contract.department = data['department']
    contract.payment_terms = get_payload_value(data, 'payment_terms')
    contract.original_contract_no = get_payload_value(data, 'original_contract_no')
    contract.original_contract_name = get_payload_value(data, 'original_contract_name')
    contract.remarks = get_payload_value(data, 'remarks')
    contract.executive_partner = get_payload_value(data, 'executive_partner')
    contract.filler = get_payload_value(data, 'filler')

    db.session.commit()

    return jsonify({'success': True})


# 作废合同
@app.route('/api/contracts/<int:contract_id>/void', methods=['POST'])
def void_contract(contract_id):
    contract = Contract.query.get_or_404(contract_id)
    contract.status = 'invalid'
    db.session.commit()
    return jsonify({'success': True})


# 检查是否可以删除
@app.route('/api/contracts/<int:contract_id>/check_delete', methods=['GET'])
def check_delete(contract_id):
    contract = Contract.query.get_or_404(contract_id)
    
    # 检查该合同编号之后是否已有其他项目被创建
    # 逻辑：查询同类型同平台下，ID比当前大的合同
    subsequent = Contract.query.filter(
        Contract.contract_type == contract.contract_type,
        Contract.platform == contract.platform,
        Contract.id > contract.id
    ).first()
    
    if subsequent:
        return jsonify({
            'can_delete': False,
            'message': '该合同编号后续已有项目创建，不可删除。如需停用，请使用“作废”功能。'
        })
    else:
        return jsonify({'can_delete': True})


# 删除合同
@app.route('/api/contracts/<int:contract_id>', methods=['DELETE'])
def delete_contract(contract_id):
    contract = Contract.query.get_or_404(contract_id)
    
    # 再次检查安全性
    subsequent = Contract.query.filter(
        Contract.contract_type == contract.contract_type,
        Contract.platform == contract.platform,
        Contract.id > contract.id
    ).first()
    
    if subsequent:
        return jsonify({'success': False, 'message': '该合同编号后续已有项目创建，不可删除。'}), 400
        
    db.session.delete(contract)
    db.session.commit()

    return jsonify({'success': True})


# 导出Excel
@app.route('/api/export/excel', methods=['GET'])
def export_excel():
    query = apply_contract_filters(Contract.query, request.args).order_by(Contract.created_at.desc(), Contract.id.desc())
    total_count = query.count()

    # 使用 write_only 模式按批次写入，避免全量工作簿常驻内存
    wb = openpyxl.Workbook(write_only=True)
    ws = wb.create_sheet(title='合同列表')

    # 设置标题样式
    header_fill = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal='center', vertical='center')

    # 表头
    headers = [
        '序号', '合同编号', '合同名称', '项目号', '合同类型', '所属平台',
        '合同金额 (元)', '签订日期', '单位名称', '企业负责人', '联系电话',
        '执行合伙人', '填表人', '所属部门', '支付条件',
        '原合同编号', '原合同名称', '状态', '备注', '创建时间', '更新时间'
    ]

    header_row = []
    for header in headers:
        cell = WriteOnlyCell(ws, value=header)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        header_row.append(cell)
    ws.append(header_row)

    # 数据行
    for index, contract in enumerate(query.yield_per(EXPORT_BATCH_SIZE), start=1):
        ws.append([
            total_count - index + 1,
            contract.contract_no,
            contract.contract_name,
            contract.project_no,
            contract.contract_type,
            contract.platform,
            float(contract.contract_amount) if contract.contract_amount else 0,
            contract.sign_date.strftime('%Y-%m-%d') if contract.sign_date else '',
            contract.company_name,
            contract.corporate_principal,
            contract.contact_phone,
            contract.executive_partner,
            contract.filler,
            contract.department,
            contract.payment_terms,
            contract.original_contract_no,
            contract.original_contract_name,
            '已作废' if contract.status == 'invalid' else '正常',
            contract.remarks,
            contract.created_at.strftime('%Y-%m-%d %H:%M:%S') if contract.created_at else '',
            contract.updated_at.strftime('%Y-%m-%d %H:%M:%S') if contract.updated_at else ''
        ])

    tmp_file = tempfile.NamedTemporaryFile(delete=False, suffix='.xlsx')
    temp_path = tmp_file.name
    tmp_file.close()

    try:
        wb.save(temp_path)
    except Exception:
        if os.path.exists(temp_path):
            os.remove(temp_path)
        raise
    finally:
        wb.close()

    @after_this_request
    def cleanup_export_file(response):
        try:
            if os.path.exists(temp_path):
                os.remove(temp_path)
        except OSError:
            pass
        return response

    return send_file(
        temp_path,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'合同列表_{datetime.now().strftime("%Y%m%d")}.xlsx',
        max_age=0
    )


@app.teardown_appcontext
def shutdown_session(exception=None):
    db.session.remove()


if __name__ == '__main__':
    with app.app_context():
        db.create_all()
    app.run(
        debug=get_bool_env('FLASK_DEBUG', False),
        host=os.getenv('HOST', '0.0.0.0'),
        port=get_int_env('PORT', 5600, minimum=1, maximum=65535)
    )
